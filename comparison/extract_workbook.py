#!/usr/bin/env python3
"""
ONE-TIME extractor: hand-maintained Excel workbook -> versioned TSVs.

This script is *build-time only*. It is committed so the derivation of
`prior_studies.tsv` and `this_study.tsv` from the legacy workbook is auditable,
but nothing in the runtime path (`plot_comparison.py`) ever opens the .xlsx.
Re-run it only if the workbook is corrected; then commit the regenerated TSVs.

Source workbook (not in git; lives under the gitignored `ignore/` tree):
    ignore/2026-04-30 Low Input Comparison/Manuscript/
        2026-04-29 Low-Input Comparison Calculations.xlsx

Sheets read
-----------
  'Per Read Comparison'          plotted reads/fg per prior-study sample
  'Per Base Comparison'          plotted bases/fg per prior-study sample
  'Mojarro et al. 2019'          Mojarro first-principles + literal counts
  'Zorzano et al. 2025'          Zorzano Supplementary Table 2, extended
  'B. Raghavendra et al. 2023'   Raghavendra Tables 2/3 + our Kraken2 reanalysis
  'S1'                           this study, Round 1 per-replicate
  'S2'                           this study, Round 2 aggregate

Conventions applied to every emitted row
----------------------------------------
    reads_per_fg = reads / (dna_pg * 1000)
    bases_per_fg = bases / (dna_pg * 1000)

Known defects in the workbook are NOT silently repaired here; they are made
explicit through the `classifier`, `source`, `verified` and `provenance_note`
columns. See README.md.
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

HERE = Path(__file__).resolve().parent
REPO = HERE.parent
DEFAULT_XLSX = (REPO / "ignore" / "2026-04-30 Low Input Comparison" / "Manuscript"
                / "2026-04-29 Low-Input Comparison Calculations.xlsx")

# Column order for both TSVs. `this_study.tsv` appends a few extra columns at
# the end; the shared prefix is what `comparison_data.py` relies on.
SHARED_COLUMNS = [
    "study",
    "study_short",
    "condition",
    "organism",
    "replicate_idx",
    "replicate_label",
    "row_type",
    "classifier",
    "reads",
    "bases",
    "dna_pg",
    "dna_basis",
    "reads_per_fg",
    "bases_per_fg",
    "verified",
    "source",
    "source_detail",
    "provenance_note",
]
THIS_STUDY_COLUMNS = SHARED_COLUMNS + ["round", "experiment", "sample_id", "mode"]

WB = "workbook '2026-04-29 Low-Input Comparison Calculations.xlsx'"


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------
def fmt(v):
    """Round-trip-exact text for a number; empty string for None."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, int):
        return str(v)
    if isinstance(v, float):
        if v.is_integer() and abs(v) < 1e15:
            return str(int(v))
        return repr(v)
    return str(v)


def per_fg(count, dna_pg):
    """reads/fg or bases/fg under the module-wide convention."""
    if count is None or dna_pg is None:
        return None
    return count / (dna_pg * 1000.0)


def write_tsv(path: Path, columns, rows):
    with path.open("w", encoding="utf-8") as fh:
        fh.write("\t".join(columns) + "\n")
        for r in rows:
            fh.write("\t".join(fmt(r.get(c)) for c in columns) + "\n")
    print(f"[wrote] {path}  ({len(rows)} rows)")


def check(label, got, want, tol=1e-9):
    """Assert an emitted value reproduces the workbook's own computed value."""
    if want is None:
        return
    if got is None:
        raise SystemExit(f"[FAIL] {label}: computed None, workbook had {want}")
    if abs(got - want) > tol * max(1.0, abs(want)):
        raise SystemExit(f"[FAIL] {label}: computed {got!r} != workbook {want!r}")


# ---------------------------------------------------------------------------
# Mojarro et al. 2019
# ---------------------------------------------------------------------------
def extract_mojarro(wb):
    ws = wb["Mojarro et al. 2019"]
    val = {ws.cell(r, 1).value: ws.cell(r, 2).value for r in range(1, ws.max_row + 1)}
    dna_pg = float(val["Mass of DNA (actual)"])          # B13 = 2 pg
    reads = int(val["Reads obtained"])                   # B14 = 5   <-- UNSOURCED
    bases = int(val["Bases obtained"])                   # B15 = 5270 <-- UNSOURCED
    est_pg = float(val["Mass of DNA (estimated)"])       # B12 first-principles

    rpf, bpf = per_fg(reads, dna_pg), per_fg(bases, dna_pg)
    check("Mojarro reads/fg", rpf, val["Reads/fg"])
    check("Mojarro bases/fg", bpf, val["Bases/fg"])

    note = (
        "DEFECT (b) UNVERIFIED / CITATION NEEDED: 'Reads obtained'=5 and "
        "'Bases obtained'=5270 are hand-entered literals in the workbook with no "
        "table, figure, page or URL recorded anywhere in the file. Only the DNA "
        "mass (2 pg) carries a source attribution ('Mojarro et al. (2019)'). "
        f"A first-principles estimate on the same sheet gives {est_pg:.4g} pg from "
        "10,000 B. spizizenii cells x 4.4276 fg/cell x 5% simulated extraction "
        "efficiency, which corroborates the mass but not the read/base counts. "
        "Do not publish this point until the counts are traced to the paper."
    )
    return [{
        "study": "Mojarro et al. 2019",
        "study_short": "Mojarro et al.",
        "condition": "2 pg Bacillus spizizenii DNA",
        "organism": "Bacillus spizizenii",
        "replicate_idx": 0,
        "replicate_label": "n/a",
        "row_type": "sample",
        "classifier": "published_table1",
        "reads": reads,
        "bases": bases,
        "dna_pg": dna_pg,
        "dna_basis": "stated_input_mass",
        "reads_per_fg": rpf,
        "bases_per_fg": bpf,
        "verified": False,
        "source": "unsourced_literal",
        "source_detail": f"{WB} sheet 'Mojarro et al. 2019' cells B13/B14/B15",
        "provenance_note": note,
    }]


# ---------------------------------------------------------------------------
# Zorzano et al. 2025
# ---------------------------------------------------------------------------
def extract_zorzano(wb):
    ws = wb["Zorzano et al. 2025"]
    rows = []
    for r in range(4, 13):                       # data rows A4:S12
        cond = ws.cell(r, 1).value
        total_bases = int(ws.cell(r, 5).value)   # E  published total bases
        total_reads = int(ws.cell(r, 6).value)   # F  published total reads
        rep_hits = int(ws.cell(r, 7).value)      # G  published SqueezeMeta hits
        k2_hits = int(ws.cell(r, 8).value)       # H  our Kraken2 minQ1 hits
        k2_bases = int(ws.cell(r, 9).value)      # I  our Kraken2 minQ1 hit bases
        basis = ws.cell(r, 14).value             # N  Measured | Calculated
        dna_fg = float(ws.cell(r, 15).value)     # O  extracted DNA, fg
        dna_pg = dna_fg / 1000.0
        # replicate index within (study, condition); Zorzano reports one pooled
        # sample per condition, so this is always 0.
        idx = 0

        # workbook's own per-fg cells, used to verify our arithmetic
        wb_hits_per_fg = ws.cell(r, 17).value    # Q = G/O  (published hits)
        wb_hitbases_per_fg = ws.cell(r, 18).value  # R = I/O (Kraken2 bases)

        dna_basis = ("measured_qubit_total_of_replicates" if basis == "Measured"
                     else "back_calculated_from_1.046_Mbp_per_ng")
        dna_note = (
            "DNA mass is the sum of all replicate Qubit readings in Zorzano "
            "Supplementary Table 1." if basis == "Measured" else
            "DNA mass was below the fluorometer detection limit; the workbook "
            "back-calculates it from the paper's main-text yield of 1.046 Mbp "
            "per ng extracted DNA (dna_ng = total_bases / 1.046e6)."
        )

        # --- variant 1: published SqueezeMeta hits on BOTH axes ---------------
        mean_len = total_bases / total_reads
        est_bases = int(round(rep_hits * mean_len))
        rpf = per_fg(rep_hits, dna_pg)
        check(f"Zorzano {cond} published reads/fg", rpf, wb_hits_per_fg)
        rows.append({
            "study": "Zorzano et al. 2025",
            "study_short": "Zorzano et al.",
            "condition": cond,
            "organism": "metagenome",
            "replicate_idx": idx,
            "replicate_label": "pooled (per 1.5 g)",
            "row_type": "sample",
            "classifier": "published_squeezemeta",
            "reads": rep_hits,
            "bases": est_bases,
            "dna_pg": dna_pg,
            "dna_basis": dna_basis,
            "reads_per_fg": rpf,
            "bases_per_fg": per_fg(est_bases, dna_pg),
            "verified": True,
            "source": "published_table",
            "source_detail": (f"{WB} sheet 'Zorzano et al. 2025' G{r} (hits), "
                              f"E{r}/F{r} (total bases/reads), O{r} (DNA fg); "
                              "adapted from Zorzano et al. 2025 Supplementary Tables 1-2"),
            "provenance_note": (
                "Reads are the paper's own SqueezeMeta-classified 'Reported Hits'. "
                "The paper does NOT publish hit bases, so bases here are a DERIVED "
                f"ESTIMATE: hits x mean read length ({mean_len:.1f} bp = published "
                "total bases / published total reads), i.e. hits are assumed to have "
                "the same length distribution as non-hits. That assumption is known "
                "to be wrong (our Kraken2 minQ1 hits average ~5-9x the overall mean "
                f"read length), so treat this bases value as an upper-bound sketch. {dna_note}"
            ),
        })

        # --- variant 2: our Kraken2 minQ1 reanalysis on BOTH axes -------------
        rows.append({
            "study": "Zorzano et al. 2025",
            "study_short": "Zorzano et al.",
            "condition": cond,
            "organism": "metagenome",
            "replicate_idx": idx,
            "replicate_label": "pooled (per 1.5 g)",
            "row_type": "sample",
            "classifier": "kraken2_q1",
            "reads": k2_hits,
            "bases": k2_bases,
            "dna_pg": dna_pg,
            "dna_basis": dna_basis,
            "reads_per_fg": per_fg(k2_hits, dna_pg),
            "bases_per_fg": per_fg(k2_bases, dna_pg),
            "verified": True,
            "source": "kraken2_reanalysis",
            "source_detail": (f"{WB} sheet 'Zorzano et al. 2025' H{r} (hits), I{r} "
                              f"(hit bases), O{r} (DNA fg); Zorzano et al. 2025 raw "
                              "reads reanalysed with wf-metagenomics v2.14.1, kraken2, "
                              "PlusPF-8 database, minQ 1"),
            "provenance_note": (
                "Both axes come from the same Kraken2 minQ1 reanalysis of the "
                "published raw reads, so reads/fg and bases/fg are internally "
                f"consistent. This is the plot default. {dna_note}"
            ),
        })
        check(f"Zorzano {cond} kraken2 bases/fg",
              per_fg(k2_bases, dna_pg), wb_hitbases_per_fg)

        # --- variant 3: the workbook's mismatched hybrid (audit only) ---------
        rows.append({
            "study": "Zorzano et al. 2025",
            "study_short": "Zorzano et al.",
            "condition": cond,
            "organism": "metagenome",
            "replicate_idx": idx,
            "replicate_label": "pooled (per 1.5 g)",
            "row_type": "sample",
            "classifier": "legacy_hybrid_workbook",
            "reads": rep_hits,
            "bases": k2_bases,
            "dna_pg": dna_pg,
            "dna_basis": dna_basis,
            "reads_per_fg": per_fg(rep_hits, dna_pg),
            "bases_per_fg": per_fg(k2_bases, dna_pg),
            "verified": True,
            "source": "legacy_spreadsheet",
            "source_detail": (f"{WB} 'Per Read Comparison'!E (= 'Zorzano et al. 2025'!Q, "
                              f"from G{r}) paired with 'Per Base Comparison'!E "
                              f"(= 'Zorzano et al. 2025'!R, from I{r})"),
            "provenance_note": (
                "DEFECT (a) AXIS INCONSISTENCY, reproduced verbatim for audit: the "
                "workbook figure takes the x value from the paper's SqueezeMeta "
                "'Reported Hits' (col G) and the y value from OUR Kraken2 minQ1 hit "
                f"bases (col I). Different classifiers: here {rep_hits} reported hits "
                f"vs {k2_hits} Kraken2 hits. Provided only so the legacy figure can be "
                "reproduced; do not use for publication."
            ),
        })

    # --- aggregate rows carried in 'Per Base Comparison' (never plotted) ------
    pb = wb["Per Base Comparison"]
    total_k2_bases = int(pb.cell(7, 4).value)
    total_k2_bases_per_fg = float(pb.cell(7, 5).value)
    total_dna_fg = sum(float(ws.cell(r, 15).value) for r in range(4, 13))
    check("Zorzano aggregate bases/fg",
          total_k2_bases / total_dna_fg, total_k2_bases_per_fg)

    rows.append({
        "study": "Zorzano et al. 2025",
        "study_short": "Zorzano et al.",
        "condition": "All conditions (average reported value)",
        "organism": "metagenome",
        "replicate_idx": 0,
        "replicate_label": "aggregate",
        "row_type": "aggregate",
        "classifier": "published_total_bases",
        "reads": None,
        "bases": None,
        "dna_pg": None,
        "dna_basis": "n/a",
        "reads_per_fg": None,
        "bases_per_fg": 1.046,
        "verified": True,
        "source": "published_table",
        "source_detail": (f"{WB} 'Per Base Comparison'!E6; Zorzano et al. 2025 main "
                          "text, 1.046 Mbp per ng extracted DNA"),
        "provenance_note": (
            "Aggregate, not a sample. The paper's headline yield of 1.046 Mbp/ng "
            "== 1.046 bases/fg, over ALL basecalled bases including unclassified "
            "and very low quality reads. Never plotted; kept because it is the most "
            "conservative published comparator (this study is ~85x higher than it)."
        ),
    })
    rows.append({
        "study": "Zorzano et al. 2025",
        "study_short": "Zorzano et al.",
        "condition": "All conditions (Kraken2 minQ1 hits, summed)",
        "organism": "metagenome",
        "replicate_idx": 1,
        "replicate_label": "aggregate",
        "row_type": "aggregate",
        "classifier": "kraken2_q1",
        "reads": None,
        "bases": total_k2_bases,
        "dna_pg": total_dna_fg / 1000.0,
        "dna_basis": "sum_of_per_condition_dna",
        "reads_per_fg": None,
        "bases_per_fg": total_k2_bases / total_dna_fg,
        "verified": True,
        "source": "kraken2_reanalysis",
        "source_detail": f"{WB} 'Per Base Comparison'!D7:E7 = SUM of 'Zorzano et al. 2025'!I4:I12",
        "provenance_note": (
            "Aggregate, not a sample; never plotted. NOTE the workbook labels this "
            "row 'hits, assuming same length as non-hits', but the formula is "
            "SUM('Zorzano et al. 2025'!I4:I12) and column I is the measured Kraken2 "
            "minQ1 hit-base count, not a read-length extrapolation. The label is "
            "stale; the number is a direct sum of measured Kraken2 bases."
        ),
    })
    return rows


# ---------------------------------------------------------------------------
# Basapathi Raghavendra et al. 2023
# ---------------------------------------------------------------------------
def extract_raghavendra(wb):
    ws = wb["B. Raghavendra et al. 2023"]
    rows = []
    seen = {}
    for r in range(3, 17):                        # data rows A3:K16
        cond = ws.cell(r, 1).value
        replicate = ws.cell(r, 2).value           # B  published replicate number
        total_reads = int(ws.cell(r, 3).value)    # C  published total reads
        species = ws.cell(r, 4).value             # D
        reported = ws.cell(r, 5).value            # E  published pass reads
        table = ws.cell(r, 6).value               # F  Table 2 / Table 3
        fastq = ws.cell(r, 7).value               # G  zenodo filename
        k2_reads = int(ws.cell(r, 8).value)       # H  our Kraken2 minQ10 pass reads
        k2_bases = int(ws.cell(r, 9).value)       # I  our Kraken2 minQ10 pass bases
        dna_pg = float(ws.cell(r, 10).value)      # J
        wb_reads_per_fg = ws.cell(r, 11).value    # K = H/(J*1000)

        idx = seen.get(cond, 0)
        seen[cond] = idx + 1
        rep_label = f"rep {replicate} / {species}"
        typo = (r == 16)   # footnote [3]: reported 263 == the file's total read count

        rpf = per_fg(k2_reads, dna_pg)
        check(f"Raghavendra row {r} reads/fg", rpf, wb_reads_per_fg)
        rows.append({
            "study": "B. Raghavendra et al. 2023",
            "study_short": "Basapathi Raghavendra et al.",
            "condition": cond,
            "organism": species,
            "replicate_idx": idx,
            "replicate_label": rep_label,
            "row_type": "sample",
            "classifier": "kraken2_q10",
            "reads": k2_reads,
            "bases": k2_bases,
            "dna_pg": dna_pg,
            "dna_basis": "stated_input_mass",
            "reads_per_fg": rpf,
            "bases_per_fg": per_fg(k2_bases, dna_pg),
            "verified": True,
            "source": "kraken2_reanalysis",
            "source_detail": (f"{WB} sheet 'B. Raghavendra et al. 2023' H{r}/I{r}/J{r}; "
                              f"fastq_pass file {fastq} from https://zenodo.org/records/8208597 "
                              "reanalysed with wf-metagenomics v2.14.1, kraken2, PlusPF-8, minQ 10"),
            "provenance_note": (
                "DEFECT (c): our Kraken2 minQ10 reanalysis of the authors' own "
                f"deposited reads gives {k2_reads} pass reads where the paper's "
                f"{table} reports {reported}. Both are kept in this file; the plot "
                "uses this reanalysis by default because it is the only variant that "
                "also yields base counts. Sample total reads in the deposited file: "
                f"{total_reads}."
                + (" The paper's value of 263 appears to be a transcription error: "
                   "263 is the TOTAL number of reads in that file, not the number "
                   "assigned to the target organism." if typo else "")
            ),
        })
        rows.append({
            "study": "B. Raghavendra et al. 2023",
            "study_short": "Basapathi Raghavendra et al.",
            "condition": cond,
            "organism": species,
            "replicate_idx": idx,
            "replicate_label": rep_label,
            "row_type": "sample",
            "classifier": "published",
            "reads": int(reported),
            "bases": None,
            "dna_pg": dna_pg,
            "dna_basis": "stated_input_mass",
            "reads_per_fg": per_fg(int(reported), dna_pg),
            "bases_per_fg": None,
            "verified": True,
            "source": "published_table",
            "source_detail": (f"{WB} sheet 'B. Raghavendra et al. 2023' E{r} "
                              f"(= Basapathi Raghavendra et al. 2023 {table}), J{r}"),
            "provenance_note": (
                "DEFECT (c), published side: the paper reports pass READS only, no "
                "base counts, so bases/fg is undefined and these rows cannot be "
                f"plotted on the y axis. Published {reported} vs our Kraken2 minQ10 "
                f"{k2_reads} for the same file."
                + (" The paper's 263 is almost certainly a typo for the file's total "
                   "read count." if typo else "")
            ),
        })
    return rows


# ---------------------------------------------------------------------------
# This study (legacy seed values)
# ---------------------------------------------------------------------------
def extract_this_study(wb):
    rows = []

    ws = wb["S1"]
    header = [c.value for c in ws[1]]
    col = {h: i for i, h in enumerate(header)}
    idx = 0
    for r in range(2, ws.max_row + 1):
        vals = [c.value for c in ws[r]]
        if vals[col["Plot"]] != 1:
            continue
        rep = int(vals[col["Replicate"]])
        dna_ng = float(vals[col["Library_DNA_ng"]])
        reads = int(vals[col["Reads"]])
        bases = int(vals[col["Bases"]])
        dna_pg = dna_ng * 1000.0
        rpf, bpf = per_fg(reads, dna_pg), per_fg(bases, dna_pg)
        check(f"S1 rep{rep} reads/fg", rpf, float(vals[col["Reads_per_fg"]]))
        check(f"S1 rep{rep} bases/fg", bpf, float(vals[col["Bases_per_fg"]]))
        rows.append({
            "study": "This study",
            "study_short": "This study",
            "condition": "D6311 community DNA + lambda carrier, depletion-mode adaptive sampling",
            "organism": "All organisms",
            "replicate_idx": idx,
            "replicate_label": f"replicate {rep}",
            "row_type": "sample",
            "classifier": "minimap2_competitive",
            "reads": reads,
            "bases": bases,
            "dna_pg": dna_pg,
            "dna_basis": "qubit_library_input",
            "reads_per_fg": rpf,
            "bases_per_fg": bpf,
            "verified": True,
            "source": "legacy_spreadsheet",
            "source_detail": f"{WB} sheet 'S1', row with Plot==1, Replicate=={rep}",
            "provenance_note": (
                "Round 1 headline ('All organisms') value as previously reported. "
                "Superseded automatically by results/<sample_id>/<mode>/"
                "<sample_id>.metrics.tsv when --results-dir is supplied."
            ),
            "round": "Round 1",
            "experiment": "lowinput_s1",
            "sample_id": f"lowinput_s1_r{rep}",
            "mode": "competitive",
        })
        idx += 1

    s2 = wb["S2"]
    lbl = {s2.cell(r, 1).value: r for r in range(1, s2.max_row + 1)}
    reads = int(s2.cell(lbl["Reads"], 2).value)
    bases = int(s2.cell(lbl["Bases"], 2).value)
    dna_ng = float(s2.cell(lbl["Remaining DNA"], 2).value)   # DNA into library prep
    dna_pg = dna_ng * 1000.0
    rpf, bpf = per_fg(reads, dna_pg), per_fg(bases, dna_pg)
    check("S2 reads/fg", rpf, float(s2.cell(lbl["Reads per fg DNA into library prep"], 2).value))
    check("S2 bases/fg", bpf, float(s2.cell(lbl["Bases per fg DNA into library prep"], 2).value))
    rows.append({
        "study": "This study",
        "study_short": "This study",
        "condition": "D6321 whole cells, post-extraction, depletion-mode adaptive sampling",
        "organism": "All organisms",
        "replicate_idx": 0,
        "replicate_label": "aggregate (3 preps pooled into one extraction)",
        "row_type": "sample",
        "classifier": "minimap2_competitive",
        "reads": reads,
        "bases": bases,
        "dna_pg": dna_pg,
        "dna_basis": "qubit_library_input",
        "reads_per_fg": rpf,
        "bases_per_fg": bpf,
        "verified": True,
        "source": "legacy_spreadsheet",
        "source_detail": f"{WB} sheet 'S2' B24/B31 (counts), B16 (0.2232 ng into library prep)",
        "provenance_note": (
            "Round 2 headline value as previously reported: 3 D6321 preps pooled "
            "into one extraction, 10 uL elution, 2 uL to Qubit, 8 uL x 0.0279 ng/uL "
            "= 0.2232 ng into library prep. NOTE this does not correspond 1:1 to any "
            "row of assets/samplesheets/lowinput_s2.csv (which lists r1/r2/r3 at "
            "0.26/0.32/0.40 ng); live pipeline output supersedes it per replicate."
        ),
        "round": "Round 2",
        "experiment": "lowinput_s2",
        "sample_id": "lowinput_s2",
        "mode": "competitive",
    })
    return rows


# ---------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--workbook", type=Path, default=DEFAULT_XLSX,
                    help="path to the legacy .xlsx (default: %(default)s)")
    ap.add_argument("--outdir", type=Path, default=HERE,
                    help="directory to write prior_studies.tsv / this_study.tsv "
                         "(default: the comparison/ module directory)")
    args = ap.parse_args()

    if not args.workbook.exists():
        sys.exit(f"[error] workbook not found: {args.workbook}\n"
                 "        This extractor is build-time only. The committed TSVs are "
                 "sufficient to regenerate the figure; see README.md.")

    import openpyxl
    wb = openpyxl.load_workbook(args.workbook, data_only=True)

    prior = (extract_mojarro(wb) + extract_zorzano(wb) + extract_raghavendra(wb))
    mine = extract_this_study(wb)

    args.outdir.mkdir(parents=True, exist_ok=True)
    write_tsv(args.outdir / "prior_studies.tsv", SHARED_COLUMNS, prior)
    write_tsv(args.outdir / "this_study.tsv", THIS_STUDY_COLUMNS, mine)
    print("[ok] all emitted per-fg values reproduce the workbook's own cells")


if __name__ == "__main__":
    main()
